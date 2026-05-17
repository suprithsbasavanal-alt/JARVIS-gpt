"""
utils/real_files.py  —  JARVIS 3.0 REAL FILE SYSTEM ACCESS
Provides genuine read/write/search access to the Mac file system.
POLICY: Never describe what a file might contain.
         Always read the actual file and report real content.
         If a file does not exist, say so explicitly.
"""

import os
import subprocess
import logging
import json
import hashlib
from pathlib import Path
from datetime import datetime

logger = logging.getLogger("JARVIS.real_files")

# File size limit for reading (avoid loading 1 GB files into RAM)
MAX_READ_BYTES = 50_000   # 50 KB — enough to summarise most text files


# ─── TEXT FILE READING ────────────────────────────────────────────────────────

def read_text_file(path: str) -> dict:
    """
    Reads a plain text / code file and returns real content.
    Returns: {ok, content, size_bytes, modified, error}
    """
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"File does not exist: {expanded}", "content": ""}
    if not os.path.isfile(expanded):
        return {"ok": False, "error": f"Path is a directory, not a file: {expanded}", "content": ""}
    try:
        size = os.path.getsize(expanded)
        modified = datetime.fromtimestamp(os.path.getmtime(expanded)).strftime("%Y-%m-%d %H:%M")
        with open(expanded, "r", errors="replace") as f:
            content = f.read(MAX_READ_BYTES)
        truncated = size > MAX_READ_BYTES
        return {
            "ok"       : True,
            "content"  : content,
            "size_bytes": size,
            "modified" : modified,
            "truncated": truncated,
            "path"     : expanded,
            "error"    : "",
        }
    except PermissionError:
        return {"ok": False, "error": f"Permission denied: {expanded}. Enable Full Disk Access in System Preferences → Privacy.", "content": ""}
    except Exception as e:
        return {"ok": False, "error": f"Read error: {e}", "content": ""}


def read_pdf(path: str) -> dict:
    """
    Reads a real PDF file using PyMuPDF (fitz).
    Returns actual text extracted from the PDF pages.
    """
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"PDF does not exist: {expanded}", "content": ""}
    try:
        import fitz   # PyMuPDF
        doc   = fitz.open(expanded)
        pages = []
        for i, page in enumerate(doc):
            if i >= 20:   # limit to first 20 pages
                pages.append("[... truncated at page 20 ...]")
                break
            pages.append(f"[Page {i+1}]\n{page.get_text()}")
        doc.close()
        full_text = "\n".join(pages)[:MAX_READ_BYTES]
        return {"ok": True, "content": full_text, "pages": len(doc), "path": expanded, "error": ""}
    except ImportError:
        return {"ok": False, "error": "PyMuPDF not installed. Run: pip install PyMuPDF", "content": ""}
    except Exception as e:
        return {"ok": False, "error": f"PDF read error: {e}", "content": ""}


def read_word_doc(path: str) -> dict:
    """Reads a real .docx Word document using python-docx."""
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"File does not exist: {expanded}", "content": ""}
    try:
        from docx import Document
        doc       = Document(expanded)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        content    = "\n".join(paragraphs)[:MAX_READ_BYTES]
        return {"ok": True, "content": content, "path": expanded, "error": ""}
    except ImportError:
        return {"ok": False, "error": "python-docx not installed. Run: pip install python-docx", "content": ""}
    except Exception as e:
        return {"ok": False, "error": f"Word read error: {e}", "content": ""}


def read_excel(path: str) -> dict:
    """
    Reads a real .xlsx Excel file using openpyxl.
    Returns actual cell values from the first sheet.
    """
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"File does not exist: {expanded}", "content": ""}
    try:
        import openpyxl
        wb      = openpyxl.load_workbook(expanded, data_only=True)
        ws      = wb.active
        rows    = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 100:   # limit to first 100 rows
                rows.append("[... truncated at 100 rows ...]")
                break
            rows.append("\t".join([str(c) if c is not None else "" for c in row]))
        content = "\n".join(rows)
        return {"ok": True, "content": content, "sheets": wb.sheetnames, "path": expanded, "error": ""}
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl", "content": ""}
    except Exception as e:
        return {"ok": False, "error": f"Excel read error: {e}", "content": ""}


def read_csv(path: str) -> dict:
    """Reads a real CSV file and returns actual data."""
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"File does not exist: {expanded}", "content": ""}
    try:
        import csv
        rows = []
        with open(expanded, "r", errors="replace") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 200:
                    rows.append("[... truncated at 200 rows ...]")
                    break
                rows.append(", ".join(row))
        return {"ok": True, "content": "\n".join(rows), "path": expanded, "error": ""}
    except Exception as e:
        return {"ok": False, "error": f"CSV read error: {e}", "content": ""}


def read_file(path: str) -> dict:
    """
    Smart file reader — detects type by extension and calls the right reader.
    Never guesses content — always reads real bytes.
    """
    ext = Path(path).suffix.lower()
    if ext == ".pdf":
        return read_pdf(path)
    elif ext in (".docx", ".doc"):
        return read_word_doc(path)
    elif ext in (".xlsx", ".xls"):
        return read_excel(path)
    elif ext == ".csv":
        return read_csv(path)
    else:
        # Default: read as text (handles .py, .js, .txt, .md, .json, .sh, etc.)
        return read_text_file(path)


# ─── FILE SEARCH ──────────────────────────────────────────────────────────────

def spotlight_search(query: str, max_results: int = 10) -> list[str]:
    """
    Uses macOS Spotlight (mdfind) to search the entire file system.
    Returns real file paths — never invented ones.
    """
    try:
        result = subprocess.run(
            ["mdfind", "-name", query],
            capture_output=True, text=True, timeout=10
        )
        paths = [p for p in result.stdout.strip().split("\n") if p]
        return paths[:max_results]
    except Exception as e:
        logger.error(f"Spotlight search failed: {e}")
        return []


def search_inside_files(keyword: str, directory: str = "~", max_results: int = 10) -> list[dict]:
    """
    Uses grep to search for a keyword inside files under the given directory.
    Returns real matches with file path and line number.
    """
    expanded = os.path.expanduser(directory)
    try:
        result = subprocess.run(
            ["grep", "-r", "-l", "-i", "--include=*.txt",
             "--include=*.py", "--include=*.md", "--include=*.js",
             keyword, expanded],
            capture_output=True, text=True, timeout=15
        )
        paths   = [p for p in result.stdout.strip().split("\n") if p]
        matches = []
        for path in paths[:max_results]:
            matches.append({"path": path, "keyword": keyword})
        return matches
    except Exception as e:
        logger.error(f"grep search failed: {e}")
        return []


# ─── FOLDER LISTING ───────────────────────────────────────────────────────────

def list_folder(path: str, max_items: int = 50) -> dict:
    """
    Lists the real contents of a folder with file sizes and dates.
    Never invents directory structure.
    """
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"Directory does not exist: {expanded}", "items": []}
    if not os.path.isdir(expanded):
        return {"ok": False, "error": f"Path is a file, not a directory: {expanded}", "items": []}
    try:
        items = []
        for entry in sorted(os.scandir(expanded), key=lambda e: e.name)[:max_items]:
            stat = entry.stat()
            items.append({
                "name"    : entry.name,
                "type"    : "dir" if entry.is_dir() else "file",
                "size"    : stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })
        return {"ok": True, "path": expanded, "items": items, "error": ""}
    except PermissionError:
        return {"ok": False, "error": f"Permission denied: {expanded}", "items": []}
    except Exception as e:
        return {"ok": False, "error": f"Listing error: {e}", "items": []}


# ─── FILE METADATA ────────────────────────────────────────────────────────────

def file_metadata(path: str) -> dict:
    """Returns real metadata for a file: size, created, modified, type."""
    expanded = os.path.expanduser(path)
    if not os.path.exists(expanded):
        return {"ok": False, "error": f"Does not exist: {expanded}"}
    try:
        stat     = os.stat(expanded)
        created  = datetime.fromtimestamp(stat.st_birthtime).strftime("%Y-%m-%d %H:%M")
        modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")
        return {
            "ok"        : True,
            "path"      : expanded,
            "size_bytes": stat.st_size,
            "size_human": _human_size(stat.st_size),
            "created"   : created,
            "modified"  : modified,
            "is_dir"    : os.path.isdir(expanded),
            "extension" : Path(expanded).suffix,
            "error"     : "",
        }
    except Exception as e:
        return {"ok": False, "error": f"Metadata error: {e}"}


def _human_size(size: int) -> str:
    """Converts bytes to human-readable string."""
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{size:.1f} PB"


# ─── FILE WRITE / CREATE ──────────────────────────────────────────────────────

def write_file(path: str, content: str) -> dict:
    """
    Writes content to a file. Creates parent directories if needed.
    Returns confirmation with the actual bytes written.
    """
    expanded = os.path.expanduser(path)
    try:
        os.makedirs(os.path.dirname(expanded), exist_ok=True)
        with open(expanded, "w") as f:
            f.write(content)
        size = os.path.getsize(expanded)
        return {"ok": True, "path": expanded, "bytes_written": size, "error": ""}
    except PermissionError:
        return {"ok": False, "error": f"Permission denied: {expanded}. Enable Full Disk Access."}
    except Exception as e:
        return {"ok": False, "error": f"Write error: {e}"}


# ─── EXECUTE REAL CODE ────────────────────────────────────────────────────────

def execute_python(code: str) -> dict:
    """
    Executes real Python code in a subprocess and returns actual stdout/stderr.
    Never fakes output — runs real code and captures real results.
    """
    try:
        result = subprocess.run(
            ["python3", "-c", code],
            capture_output=True, text=True, timeout=30
        )
        return {
            "ok"        : result.returncode == 0,
            "stdout"    : result.stdout[:3000],
            "stderr"    : result.stderr[:1000],
            "returncode": result.returncode,
            "error"     : result.stderr if result.returncode != 0 else "",
        }
    except subprocess.TimeoutExpired:
        return {"ok": False, "stdout": "", "stderr": "Execution timed out after 30 seconds.", "returncode": -1, "error": "timeout"}
    except Exception as e:
        return {"ok": False, "stdout": "", "stderr": str(e), "returncode": -1, "error": str(e)}


def execute_bash(command: str) -> dict:
    """
    Executes a real Bash command and returns actual stdout/stderr.
    NEVER shows fake terminal output.
    """
    try:
        result = subprocess.run(
            command, shell=True, capture_output=True, text=True, timeout=30
        )
        return {
            "ok"        : result.returncode == 0,
            "stdout"    : result.stdout[:3000],
            "stderr"    : result.stderr[:1000],
            "returncode": result.returncode,
            "command"   : command,
            "error"     : result.stderr if result.returncode != 0 else "",
        }
    except subprocess.TimeoutExpired:
        return {"ok": False, "stdout": "", "stderr": f"Command timed out: {command}", "returncode": -1, "error": "timeout"}
    except Exception as e:
        return {"ok": False, "stdout": "", "stderr": str(e), "returncode": -1, "error": str(e)}
